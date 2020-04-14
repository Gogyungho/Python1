import requests
from bs4 import BeautifulSoup

from openpyxl import load_workbook

# data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("mymusic.xlsx", data_only=True)
# 시트 이름으로 불러오기
load_ws = load_wb['Sheet1']

dates = ['20200316','20200317',	'20200318',	'20200319',	'20200320',	'20200321',	'20200322',	'20200323',	'20200324',	'20200325',	'20200326',	'20200327',	'20200328',	'20200329',	'20200330',	'20200331',	'20200401',	'20200402',	'20200403',	'20200404',	'20200405',	'20200406',	'20200407',	'20200408',	'20200409',	'20200410',	'20200411',	'20200412',	'20200413',	'20200414']

j = 2
for date in dates:
    
    headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
    data = requests.get('https://www.genie.co.kr/chart/top200?ditc=D&ymd=' + date + '&hh=23&rtm=N&pg=1' ,headers=headers)
    
    soup = BeautifulSoup(data.text, 'html.parser')

    trs = soup.select('#body-content > div.newest-list > div > table > tbody > tr')

    load_ws.cell(1, j, date)

    i=2 #몇번째 줄에서 시작이 되는가

    for tr in trs:
        title = tr.select_one('td.info > a.title.ellipsis').text.strip()
        artist = tr.select_one('td.info > a.artist.ellipsis').text.strip()
        rank = tr.select_one('td.number').text[0:2].strip()

        #load_ws.cell(i, 1, rank) 랭크를 매번 찍어줄 필요가 없음
        load_ws.cell(i, j, title + ' - ' +artist)
        
        i+=1
    j+=1

    load_wb.save("mymusic.xlsx")
